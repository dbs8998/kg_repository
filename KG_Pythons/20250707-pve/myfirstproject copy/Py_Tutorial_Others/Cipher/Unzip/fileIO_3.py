# 엑셀 파일 IO
# pip install openpyxl

from openpyxl import Workbook

# 1. 생성
wb = Workbook()    # workbook 객체 생성
ws = wb.active     # 시트 선택/활성화

# 데이터 입력
ws.title = "Kairos1"
ws["A1"] = '이름'
ws["B1"] = '작업팀'
ws['C1'] = '생산량'

ws.append(['이미남', '로봇1', 100])
ws.append(['김미녀', '로봇2', 120])
ws.append(['박반장', '로봇3', 140])

# 저장
file_path = r"Cipher\0_resource\test_3.xlsx"
wb.save(file_path)

print(f"엑셀 생성: {file_path}")

# 2. 읽기
from openpyxl import load_workbook

# 파일 불러오기
file_path = r"Cipher\0_resource\test_3.xlsx"
wb = load_workbook(file_path)

# 활성화된 시트 선택
ws = wb.active

# 데이터 출력
for row in ws.iter_rows(values_only=True):
    print(row)

# 파일 닫기
wb.close()

#3. 파일 업데이트
# 파일 불러오기
file_path = r"Cipher\0_resource\test_3.xlsx"
wb = load_workbook(file_path)

# 활성화된 시트 선택
ws = wb.active

# 데이터 수정
for row in ws.iter_lows(min_row=2, values_only=True):
    if row[0] == '철수':
        ws.cell(row=row[0], column=2).value = 35

# 엑셀 파일 저장
wb.save(file_path)

print(f"파일 수정: {file_path}")

wb.close()

#4. 파일 삭제

import os

if os.path.exists(file_path):
    os.remove(file_path)
    print(f"파일 {file_path}가 삭제됨")
else:
    print("파일이 존재하지 않습니다.")










with Workbook() as workbook:
    ws = workbook.active
    ws.title = "Kairos1"
    ws["A1"] = '이름'
    ws["B1"] = '작업팀'
    ws['C1'] = '생산량'
    ws['A2'] = '이미남'
    ws['B2'] = '로봇1'
    ws['C2'] = 100

print(f"엑셀생성: {file_path}")

# # 파일 읽기
# with load_workbook(filename=file_path, read_only=True, encoding='utf-8') as workbook:
#     ws = workbook.active
#     name = ws["A2"].value
#     team = ws['B2'].value
    
# print(f'이름: {name}, 소속: {team}')

# # 파일 업데이트
# newOutput = 200
# with load_workbook(filename=file_path) as workbook:
#     ws = workbook.active
#     ws["B2"] = newOutput
#     workbook.save(file_path)

# print(f"파일 갱신: {file_path}")

# # 파일 삭제
# import os

# if os.path.exists(file_path):
#     os.remove(file_path)
#     print(f"파일 {file_path}가 삭제됨")
# else:
#     print("파일이 존재하지 않습니다.")


